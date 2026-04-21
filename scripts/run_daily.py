#!/usr/bin/env python3
"""
Nok Recommerce — daily orders automation.

Runs in GitHub Actions at ~7am ET. Pulls the last 34 days of orders from
Extensiv Integration Manager, filters to a 33-day window ending yesterday,
builds a three-tab pivot workbook, and posts a Slack message + xlsx
attachment to #resale.

Environment variables required:
  IM_MERCHANT_USER, IM_MERCHANT_KEY  — Extensiv IM merchant-level credentials
  SLACK_BOT_TOKEN                    — xoxb- token with chat:write + files:write
  SLACK_CHANNEL_ID                   — channel to post to (default: C0961GBT4BY, #resale)

Exit codes:
  0  — success
  1  — failure (FAILED message is posted to Slack before exiting)
"""
from __future__ import annotations

import json
import os
import sys
import traceback
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

# Local modules (scripts/ on sys.path)
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from client import ExtensivIM, ExtensivIMError  # noqa: E402
from brand_mapper import BrandMapper  # noqa: E402

# Third-party
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from slack_sdk import WebClient  # noqa: E402
from slack_sdk.errors import SlackApiError  # noqa: E402

REPO_ROOT = HERE.parent
BRAND_MAP = REPO_ROOT / "references" / "brand_map.json"
OUT_DIR = Path(os.environ.get("OUT_DIR", REPO_ROOT / "out"))
OUT_DIR.mkdir(parents=True, exist_ok=True)

SLACK_CHANNEL = os.environ.get("SLACK_CHANNEL_ID", "C0961GBT4BY")  # #resale

# Extensiv IM order_source → friendly marketplace display name.
# Single source of truth. Add new entries here when a new marketplace appears.
MARKETPLACE = {
    "Amazon_CA2": "Amazon CA",
    "Amazon_US2": "Amazon US",
    "Amazon_US3": "Amazon US",
    "Walmart3":   "Walmart CA",
    "Walmart2":   "Walmart US",
    "Walmart":    "Walmart US",
    "Ebay_US":    "eBay US",
    "Ebay_CA":    "eBay CA",
    "Best Buy Canada": "Best Buy CA",
    "BestBuy_CA": "Best Buy CA",
    "Shopify":    "Shopify",
    "b2b":        "B2B",
    "B2B":        "B2B",
}


def mkt(src): return MARKETPLACE.get(src or "", src or "Unknown")
def units(o): return sum(int(it.get("quantity") or 0) for it in (o.get("items") or []))
def mdy(d): return f"{d.month}/{d.day}"


def parse_d(o):
    s = o.get("order_date") or (o.get("created_date_time") or "")[:10]
    try:
        y, m, d = map(int, s[:10].split("-"))
        return date(y, m, d)
    except Exception:
        return None


def post_failure(slack: WebClient, reason: str, detail: str = ""):
    text = f"*Daily orders refresh — FAILED*\n\n{reason}"
    if detail:
        text += f"\n\n{detail}"
    try:
        slack.chat_postMessage(channel=SLACK_CHANNEL, text=text)
    except SlackApiError as e:
        print(f"Failed to post failure message: {e}", file=sys.stderr)


def build_workbook(orders, brands, markets, start33, yesterday, xlsx_path: Path):
    bold = Font(bold=True)
    hdr = PatternFill("solid", fgColor="D9E1F2")

    bm_units = defaultdict(lambda: defaultdict(int))
    for o in orders:
        bm_units[o.get("_brand") or "Unmapped"][o["_mkt"]] += o["_units"]
    daily_units = defaultdict(lambda: defaultdict(int))
    for o in orders:
        daily_units[o["_d"]][o.get("_brand") or "Unmapped"] += o["_units"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Brand x Marketplace"
    ws.cell(row=1, column=1, value=f"Units — {start33} to {yesterday} (33 days)").font = Font(bold=True, size=12)
    ws.cell(row=3, column=1, value="Brand").font = bold
    for j, m in enumerate(markets, 2):
        c = ws.cell(row=3, column=j, value=m); c.font = bold; c.fill = hdr
    tot_col = len(markets) + 2
    c = ws.cell(row=3, column=tot_col, value="Total"); c.font = bold; c.fill = hdr
    for i, b in enumerate(brands, 4):
        ws.cell(row=i, column=1, value=b).font = bold
        rt = 0
        for j, m in enumerate(markets, 2):
            v = bm_units[b][m]
            ws.cell(row=i, column=j, value=v or None)
            rt += v
        ws.cell(row=i, column=tot_col, value=rt).font = bold
    trow = 4 + len(brands)
    ws.cell(row=trow, column=1, value="Total").font = bold
    gt = 0
    for j, m in enumerate(markets, 2):
        ct = sum(bm_units[b][m] for b in brands); gt += ct
        ws.cell(row=trow, column=j, value=ct).font = bold
    ws.cell(row=trow, column=tot_col, value=gt).font = bold
    ws.column_dimensions["A"].width = 18
    for j in range(2, tot_col + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13

    ws2 = wb.create_sheet("Daily by Brand")
    ws2.cell(row=1, column=1, value=f"Daily units — {start33} to {yesterday}").font = Font(bold=True, size=12)
    ws2.cell(row=3, column=1, value="Date").font = bold
    for j, b in enumerate(brands, 2):
        c = ws2.cell(row=3, column=j, value=b); c.font = bold; c.fill = hdr
    tc2 = len(brands) + 2
    c = ws2.cell(row=3, column=tc2, value="Total"); c.font = bold; c.fill = hdr
    dates = [start33 + timedelta(days=i) for i in range(33)]
    for i, d in enumerate(dates, 4):
        ws2.cell(row=i, column=1, value=d.isoformat())
        rt = 0
        for j, b in enumerate(brands, 2):
            v = daily_units[d][b]
            ws2.cell(row=i, column=j, value=v or None)
            rt += v
        ws2.cell(row=i, column=tc2, value=rt or None)
    ws2.column_dimensions["A"].width = 12
    for j in range(2, tc2 + 1):
        ws2.column_dimensions[get_column_letter(j)].width = 13

    ws3 = wb.create_sheet("Orders")
    cols = ["order_date", "marketplace", "order_source", "brand", "record_no", "cust_ref",
            "units", "grand_total", "currency", "ship_city", "ship_state", "ship_country", "order_status"]
    for j, k in enumerate(cols, 1):
        c = ws3.cell(row=1, column=j, value=k); c.font = bold; c.fill = hdr
    for i, o in enumerate(sorted(orders, key=lambda o: (o["_d"], o.get("_mkt") or ""), reverse=True), 2):
        try:
            gtv = float(o.get("grand_total") or 0)
        except Exception:
            gtv = 0
        for j, v in enumerate([
            o["_d"].isoformat() if o["_d"] else None, o["_mkt"], o.get("order_source"),
            o.get("_brand") or "Unmapped", o.get("record_no"), o.get("cust_ref"),
            o["_units"], gtv, o.get("currency_code"),
            o.get("ship_city"), o.get("ship_state"), o.get("ship_country"), o.get("order_status"),
        ], 1):
            ws3.cell(row=i, column=j, value=v)
    for j, w in enumerate([12, 14, 14, 14, 20, 14, 8, 12, 10, 16, 8, 10, 14], 1):
        ws3.column_dimensions[get_column_letter(j)].width = w
    ws3.freeze_panes = "A2"

    wb.save(xlsx_path)


def build_slack_message(orders, brands, start7, yesterday) -> str:
    y = [o for o in orders if o["_d"] == yesterday]
    yt_u = sum(o["_units"] for o in y)
    yt_o = len(y)
    by_brand = defaultdict(int)
    by_bm = defaultdict(lambda: defaultdict(int))
    for o in y:
        b = o.get("_brand") or "Unmapped"
        by_brand[b] += o["_units"]
        by_bm[b][o["_mkt"]] += o["_units"]
    w7 = [o for o in orders if start7 <= o["_d"] <= yesterday]
    w7_u = sum(o["_units"] for o in w7)
    w7_o = len(w7)

    lines = [f"*Daily orders — {mdy(yesterday)}*", ""]
    ul = "unit" if yt_u == 1 else "units"
    ol = "order" if yt_o == 1 else "orders"
    lines += [f"Total: {yt_o} {ol} · {yt_u} {ul}", ""]
    for b in brands:
        u = by_brand.get(b, 0)
        if u == 0:
            lines.append(f"• *{b}*: 0 units")
        else:
            parts = [f"{m} {c}" for m, c in sorted(by_bm[b].items(), key=lambda kv: (-kv[1], kv[0]))]
            lines.append(f"• *{b}*: {u} {'unit' if u == 1 else 'units'} ({', '.join(parts)})")
    lines += [
        "",
        f"Last 7 days: {w7_o} {'order' if w7_o == 1 else 'orders'} · {w7_u} {'unit' if w7_u == 1 else 'units'}",
    ]
    return "\n".join(lines)


def main() -> int:
    try:
        slack = WebClient(token=os.environ["SLACK_BOT_TOKEN"])
    except KeyError:
        print("SLACK_BOT_TOKEN is not set", file=sys.stderr)
        return 1

    today = date.today()
    yesterday = today - timedelta(days=1)
    start33 = yesterday - timedelta(days=32)
    start7 = yesterday - timedelta(days=6)

    # --- Step 1: pull from Extensiv IM ---
    try:
        im = ExtensivIM.from_env(level="merchant")
        from_ts = (datetime.now(timezone.utc) - timedelta(days=34)).strftime("%Y-%m-%dT00:00:00+00:00")
        to_ts = (datetime.now(timezone.utc) + timedelta(days=1)).strftime("%Y-%m-%dT00:00:00+00:00")
        orders_all = im.list_all_orders(status="any", from_date=from_ts, to_date=to_ts)
        print(f"Pulled {len(orders_all)} orders from Extensiv IM")
    except (ExtensivIMError, Exception) as e:
        detail = f"```\n{traceback.format_exc()[-1200:]}\n```"
        post_failure(slack, f"IM pull failed: {type(e).__name__}: {e}", detail)
        return 1

    if not orders_all:
        post_failure(slack, "IM returned zero orders for the 34-day window (almost certainly a bug)")
        return 1

    # --- Step 2: brand mapping + guardrail ---
    try:
        mapper = BrandMapper(BRAND_MAP)
        for o in orders_all:
            mapper.enrich_order(o)
        audit = mapper.audit(orders_all)
    except Exception as e:
        post_failure(slack, f"Brand mapping failed: {type(e).__name__}: {e}")
        return 1

    if audit["unmapped_skus"]:
        lines = []
        for u in audit["unmapped"][:15]:
            lines.append(
                f"• `{u['sku']}` — {u['units']} units — "
                f"{', '.join(u['storefronts'])} — {u['sample_description'][:60]}"
            )
        detail = (
            f"{audit['unmapped_skus']} unmapped SKU(s). "
            "Add rules to `references/brand_map.json`:\n\n" + "\n".join(lines)
        )
        post_failure(slack, "Unmapped SKUs — export halted", detail)
        return 1

    # --- Step 3: build window and pivots ---
    for o in orders_all:
        o["_d"] = parse_d(o)
        o["_mkt"] = mkt(o.get("order_source"))
        o["_units"] = units(o)
    orders = [o for o in orders_all if o["_d"] and start33 <= o["_d"] <= yesterday]

    brands = sorted({o.get("_brand") or "Unmapped" for o in orders})
    markets = sorted({o["_mkt"] for o in orders})

    xlsx_path = OUT_DIR / "nok_orders_data_dump.xlsx"
    try:
        build_workbook(orders, brands, markets, start33, yesterday, xlsx_path)
        print(f"Wrote {xlsx_path}")
    except Exception as e:
        post_failure(slack, f"Workbook build failed: {type(e).__name__}: {e}")
        return 1

    # --- Step 4: build + post Slack message with xlsx attachment ---
    body = build_slack_message(orders, brands, start7, yesterday)
    print("\n" + body + "\n")

    try:
        slack.files_upload_v2(
            channel=SLACK_CHANNEL,
            file=str(xlsx_path),
            filename=f"nok_orders_{yesterday.isoformat()}.xlsx",
            title=f"Nok orders — {yesterday.isoformat()}",
            initial_comment=body,
        )
        print("Posted to Slack (with xlsx attachment)")
    except SlackApiError as e:
        print(f"files_upload_v2 failed: {e}. Falling back to chat_postMessage.", file=sys.stderr)
        try:
            slack.chat_postMessage(channel=SLACK_CHANNEL, text=body)
            print("Posted to Slack (message only — xlsx attach failed)")
        except SlackApiError as ee:
            print(f"chat_postMessage also failed: {ee}", file=sys.stderr)
            return 1

    try:
        (OUT_DIR / "nok_daily_slack_message.txt").write_text(body)
    except Exception:
        pass

    return 0


if __name__ == "__main__":
    sys.exit(main())
